import openpyxl
import os.path

project_path = os.path.dirname(os.path.dirname(__file__))
case_path = os.path.join(project_path,"api_test\\case.xlsx")


class HandleExcel:

    def __init__(self,file_path):
        """
        初始化
        :param file_path:文件路径
        """
        self.path = file_path

    def load_excel(self):
        """
        加载表格
        """
        open_excel = openpyxl.load_workbook(self.path)
        return open_excel

    def get_sheet_data(self, index=None):
        """
        获取sheet的内容
        :param index:sheet编号
        :return: data
        """
        sheet_name = self.load_excel().sheetnames
        if index is None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, column):
        """
        获取某个单元格的内容
        :param row:行
        :param column:列
        :return:data
        """
        data = self.get_sheet_data().cell(row=row, column=column).value
        return data

    def get_rows(self, index=None):
        """
        获取行数
        :param index:sheet
        :return:行数
        """
        row = self.get_sheet_data(index).max_row
        return row

    def get_rows_value(self, row, index=None):
        """
        获取行的内容
        :param row:行
        :param index:sheet
        :return:data
        """
        row_list = []
        for i in self.get_sheet_data(index)[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, index, row, cols, value):
        """
        写入数据
        :param index: sheet编号，从0开始
        :param row:行
        :param cols:列
        :param value:数据
        """
        wb = self.load_excel()
        wr = wb.worksheets[index]
        wr.cell(row, cols, value)
        wb.save(self.path)

    def get_columns_value(self, key=None):
        """
        默认key为空，获取A列数据:所有的case编号
        :param key:列
        :return:data
        """
        column_list_data = []
        if key is None:
            key = "A"
        column_list = self.get_sheet_data()[key]
        for i in column_list:
            column_list_data.append(i.value)
        return column_list_data

    def get_row_number(self, case_id):
        """
        获取case_id行号（根据被依赖case_id获取被依赖case行号：传入case_id，拿到case_id的行号）
        :param case_id: 用例id（condition_data中的方法分离前置条件获得）
        :return: 行号（返回被依赖case的行号，从而执行case）
        """
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num += 1
        return num

    def get_excel_data(self, index=None):
        """
        获取单个sheet的excel数据
        :param index:sheet编号
        :return:data
        """
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 2, index))
        return data_list


    def get_excel_datas(self,index_list):
        """
        获取多个sheet的excel数据
        :param index_list: sheet编号列表
        :return: data
        """
        data_lists = []
        for i in index_list:
            data_lists += self.get_excel_data(i)
        return data_lists


case_data = HandleExcel(case_path)

if __name__ == '__main__':
    pass